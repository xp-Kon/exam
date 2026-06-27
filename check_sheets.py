import openpyxl
wb = openpyxl.load_workbook("Web编程题库汇总.xlsx")
print(f"Sheets: {wb.sheetnames}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet '{name}': {ws.max_row} rows x {ws.max_column} cols")
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True):
        print(f"  {[str(c)[:50] if c else None for c in row]}")
