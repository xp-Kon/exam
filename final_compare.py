# -*- coding: utf-8 -*-
import openpyxl, re, json

# Read Excel
wb = openpyxl.load_workbook("Web编程题库汇总.xlsx")
ws = wb[wb.sheetnames[0]]
rows = [cell[0].value for cell in ws.iter_rows()]

excel_qs = {}
i = 0
while i < len(rows):
    val = rows[i]
    if val and isinstance(val, str) and re.match(r'^题目\d+[：:]', str(val)):
        q_num = int(re.search(r'题目(\d+)', str(val)).group(1))
        q_text = str(rows[i+1]).strip() if i+1 < len(rows) and rows[i+1] else ''
        next_val = str(rows[i+2]).strip() if i+2 < len(rows) and rows[i+2] else ''
        if next_val.startswith('正确答案'):
            ans = re.sub(r'^正确答案[：:]\s*', '', next_val).strip()
            ans_list = [a.strip() for a in ans.split(';') if a.strip()]
            excel_qs[q_num] = {'q': q_text, 'opts': None, 'ans': ans_list, 'type': 'fill'}
            i += 4
        else:
            opts_raw = next_val
            ans_raw = str(rows[i+3]).strip() if i+3 < len(rows) and rows[i+3] else ''
            ans = re.sub(r'^正确答案[：:]\s*', '', ans_raw).strip()
            opt_vals = {}
            for om in re.finditer(r'([A-D])、\s*(.*?)(?=\s+[A-D]、|$)', opts_raw):
                opt_vals[om.group(1)] = om.group(2).strip()
            excel_qs[q_num] = {'q': q_text, 'opts': opt_vals, 'ans': [ans], 'type': 'single'}
            i += 5
    else:
        i += 1

# Read data.js JSP questions from JSON
with open('jsp_datajs.json', 'r', encoding='utf-8') as f:
    js_list = json.load(f)

js_qs = {q['id']: q for q in js_list}
print(f"Excel: {len(excel_qs)}, data.js: {len(js_qs)}")

# Compare
diffs = 0
for qid in sorted(excel_qs.keys()):
    eq = excel_qs[qid]
    jq = js_qs.get(qid)
    if not jq:
        print(f"Q#{qid}: MISSING in data.js")
        diffs += 1
        continue
    
    # Compare options
    if eq['opts']:
        jq_opts = jq.get('options')
        if not jq_opts:
            print(f"Q#{qid}: data.js NULL opts, Excel: {eq['opts']}")
            diffs += 1
            continue
        for k in ['A', 'B', 'C', 'D']:
            ev = eq['opts'].get(k, '')
            jv = jq_opts.get(k, '')
            if ev != jv:
                print(f"\nQ#{qid} OPT {k}: DIFFERS")
                print(f"  Excel:  '{ev}'")
                print(f"  data.js: '{jv}'")
                diffs += 1
    
    # Compare answers
    eq_ans = eq['ans']
    jq_ans = jq.get('answer', [])
    if eq_ans != jq_ans:
        print(f"\nQ#{qid}: ANSWER DIFFERS")
        print(f"  Excel:  {eq_ans}")
        print(f"  data.js: {jq_ans}")
        diffs += 1

print(f"\nTotal differences: {diffs}")
